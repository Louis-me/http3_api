__author__ = "shikun"
from openpyxl import load_workbook
from common.customConst import Const
url_ = []  # 申明list
workbook_ = load_workbook(filename=Const.PICT_PARAMS_RESULT)  # 导入工作表
sheetnames = workbook_.get_sheet_names()  # 获得表单名字
sheet = workbook_.get_sheet_by_name(sheetnames[0])  # 从工作表中提取某一表单
nrows = sheet.nrows
for rowNum in range(1, nrows):
    url_.append(rowNum) # 获得数据
print(url_)